from datetime import datetime, time, timedelta
from tools.reqtool import *
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def merge_duplicate_ips(ip_list):
    ip_dict = {}
    for item in ip_list:
        key = (item['private_ip'], item['public_ip'])
        if key not in ip_dict:
            item['start_usage_time'] = datetime.strptime(item['discovery_time'],
                                                         "%Y-%m-%dT%H:%M:%S.%f+00:00") + timedelta(hours=8)
            item['latest_usage_time'] = datetime.strptime(item['discovery_time'],
                                                          "%Y-%m-%dT%H:%M:%S.%f+00:00") + timedelta(hours=8)
            del item['discovery_time']
            ip_dict[key] = item
        else:
            existing_item = ip_dict[key]
            new_start_time = datetime.strptime(item['discovery_time'], "%Y-%m-%dT%H:%M:%S.%f+00:00") + timedelta(
                hours=8)
            if new_start_time < existing_item['start_usage_time']:
                existing_item['start_usage_time'] = new_start_time
            if new_start_time > existing_item['latest_usage_time']:
                existing_item['latest_usage_time'] = new_start_time

    result_list = []
    for item in ip_dict.values():
        item['start_usage_time'] = item['start_usage_time'].strftime("%Y/%m/%d %H:%M:%S")
        item['latest_usage_time'] = item['latest_usage_time'].strftime("%Y/%m/%d %H:%M:%S")
        result_list.append(item)

    return result_list


def get_time_range(dicts):
    if not dicts:
        return None
    # 初始化最大值和最小值，将第一个元素的时间作为初始值
    start_dt = end_dt = datetime.fromisoformat(dicts[0]["discoveryTime"]) + timedelta(hours=8)
    for d in dicts[1:]:
        dt = datetime.fromisoformat(d["discoveryTime"]) + timedelta(hours=8)
        if dt > end_dt:
            end_dt = dt
        elif dt < start_dt:
            start_dt = dt
    return (format_datetime(start_dt), format_datetime(end_dt))


def format_datetime(iso_value):
    # 将输入数据转换为字符串类型
    iso_string = str(iso_value)
    # 将 ISO 8601 格式的时间字符串解析为 datetime 对象
    dt = datetime.fromisoformat(iso_string)

    # 将 datetime 对象格式化为指定的字符串格式
    target_str = dt.strftime("%Y/%m/%d %H:%M:%S")

    return target_str


def get_current_date():
    now = datetime.now()
    formatted_date = now.strftime("%Y-%m-%d")
    return formatted_date


def remove_outdated_records(now_time, data_dict):
    # 复制原始字典，以免直接修改原始数据
    new_dict = data_dict.copy()

    # 获取 data 列表
    data_list = new_dict['data']
    new_data_list = []

    # 遍历每个词典并处理
    for data in data_list:
        #   if 'discoveryTime' in data and '2023-03-29' in data['discoveryTime']:
        if 'discoveryTime' in data and now_time in data['discoveryTime']:
            new_data_list.append(data)

    # 更新新字典中的数据列表
    new_dict['count'] = len(new_data_list)
    new_dict['data'] = new_data_list

    return new_dict


def save_to_excel(ip_list, template_file_path, output_file_path):
    if not ip_list:
        return

    # 加载模板文件
    wb = load_workbook(template_file_path)

    # 选择第一个工作表
    ws = wb.active

    # 写入数据到新的xlsx文件
    for i, item in enumerate(ip_list):
        row = i + 2  # 第一行已经写入了标题，因此从第二行开始写入
        for j, key in enumerate(
                ['private_ip', 'local_network', 'public_ip', 'isp', 'start_usage_time', 'latest_usage_time']):
            col = get_column_letter(j + 1)
            cell_value = str(item.get(key))
            ws.cell(row=row, column=j + 1, value=cell_value)

        # # 如果大于一行，自动在第二行下方添加一行，并插入数据
        # if row > 2:
        #     ws.insert_rows(row)

    # 保存Excel文件
    wb.save(output_file_path)


def get_table_columns_from_packet(inIpDomainName, owner_details_dict):
    table_info_list = []
    if 'count' in owner_details_dict and owner_details_dict['count'] >= 2:
        for owner_outnet_dict in owner_details_dict['data']:
            if 'inIp' in owner_outnet_dict and 'outIp' in owner_outnet_dict and 'ipInfo' in owner_outnet_dict:
                # 内部IP：private_ip
                # IP地址归属：local_network
                # 互联网出口IP地址：public_ip
                # 互联网出口地址归属：isp
                # 出口IP使用起始时间：start_usage_time
                # 出口IP使用最新时间：latest_usage_time
                dict = {}
                # 内部IP地址
                dict['private_ip'] = owner_outnet_dict['inIp']
                # 内部归属
                dict['local_network'] = inIpDomainName
                # 互联网出口IP地址
                dict['public_ip'] = owner_outnet_dict['outIp']
                # 互联网出口地址归属
                dict['isp'] = owner_outnet_dict['ipInfo']
                # 发现时间
                dict['discovery_time'] = owner_outnet_dict['discoveryTime']

                table_info_list.append(dict)

    table_info_list = merge_duplicate_ips(table_info_list)

    return table_info_list


def process_data(session, illicit_outnet_dict, now_time='2023-03-29'):
    illicit_outnet_count = []
    # 检查 count 是否为 0
    if 'count' in illicit_outnet_dict and illicit_outnet_dict['count'] == 0:
        print("count 为 0，不需要处理")
        return illicit_outnet_count

    # 获取 非法外联 列表
    illicit_outnet_list = illicit_outnet_dict['data']

    # 从非法外联列表里面获取，具体归属单位，然后请求单位违规外联详细信息
    for illicit_outnet_owner in illicit_outnet_list:
        if 'inIp' in illicit_outnet_owner \
                and 'gatewaySeries' in illicit_outnet_owner \
                and 'inIpDomainName' in illicit_outnet_owner:
            in_ip = illicit_outnet_owner['inIp']
            gateway_series = illicit_outnet_owner['gatewaySeries']
            # 内网IP地址归属
            inIpDomainName = illicit_outnet_owner['inIpDomainName']

            outnet_owner_details_dict = get_outnet_details(session=session, in_ip=in_ip,
                                                           gateway_series=gateway_series)
            # 去除过期数据，也就是今天以外的数据
            outnet_owner_details_dict = remove_outdated_records(now_time, outnet_owner_details_dict)
            # 从归属获取表格数据
            table_info_list = get_table_columns_from_packet(inIpDomainName, outnet_owner_details_dict)

            illicit_outnet_count.append(table_info_list)

    return illicit_outnet_count


def generate_filename(s, today):
    """生成文件名"""

    def get_last_word(s):
        """获取最后一个'->'分隔的字符"""
        return s.rsplit('->', 1)[-1]

    # 获取最后一个'->'分隔的字符
    last_word = get_last_word(s)

    today = today.replace("-", "")
    filename = f"违规外联-{last_word}{today}.xlsx"

    return filename
